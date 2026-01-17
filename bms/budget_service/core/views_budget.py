from rest_framework import mixins, viewsets, status
import csv
import json
from decimal import Decimal

import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from django.db import transaction
from django.db.models import Sum, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, status, viewsets, filters, serializers
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes, action
from drf_spectacular.utils import (
    extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes, OpenApiExample
)
from core.authentication import JWTCookieAuthentication

from .service_authentication import APIKeyAuthentication

from .models import (
    Account, AccountType, BudgetProposal, BudgetTransfer, Department, DepartmentBudgetCap, ExpenseCategory,
    FiscalYear, BudgetAllocation, Expense, JournalEntry, JournalEntryLine,
    ProposalComment, ProposalHistory, SubCategoryBudgetCap, UserActivityLog, Project
)
from .permissions import CanSubmitForApproval, IsTrustedService, IsBMSFinanceHead, IsBMSUser, IsBMSAdmin
from .pagination import FiveResultsSetPagination, SixResultsSetPagination, StandardResultsSetPagination
from .serializers import FiscalYearSerializer
from .views_utils import get_user_bms_role
from .serializers_budget import (
    AccountDropdownSerializer,
    AccountSetupSerializer,
    AccountTypeDropdownSerializer,
    BudgetAdjustmentSerializer,
    BudgetProposalListSerializer,
    BudgetProposalMessageSerializer,
    BudgetTransferSerializer,
    ExternalJournalEntrySerializer,
    JournalEntryDetailSerializer,
    ProposalCommentCreateSerializer,
    BudgetProposalSummarySerializer,
    BudgetProposalDetailSerializer,
    DepartmentDropdownSerializer,
    ExpenseCategoryVarianceSerializer,
    JournalEntryCreateSerializer,
    JournalEntryListSerializer,
    LedgerViewSerializer,
    ProposalCommentSerializer,
    ProposalHistorySerializer,
    ProposalReviewBudgetOverviewSerializer,
    ProposalReviewSerializer,
    SupplementalBudgetRequestSerializer
)


# --- Budget Proposal Page Views ---
# @extend_schema(
#     tags=['Budget Proposal Page'],
#     summary="List budget proposals with filters",
#     parameters=[
#         OpenApiParameter(name="department", type=str), OpenApiParameter(
#             name="status", type=str),
#         OpenApiParameter(name="search", type=str),
#     ],
#     responses={200: BudgetProposalListSerializer(many=True)}
# )
# class BudgetProposalListView(generics.ListAPIView):
#     serializer_class = BudgetProposalListSerializer
#     permission_classes = [IsAuthenticated]
#     # MODIFIED: Use new pagination class for 5 items per page
#     pagination_class = FiveResultsSetPagination

#     def get_queryset(self):  # Logic seems fine
#         queryset = BudgetProposal.objects.select_related('department').filter(
#             is_deleted=False)  # Added select_related and is_deleted filter
#         department_code = self.request.query_params.get('department')
#         status_filter = self.request.query_params.get('status')
#         search = self.request.query_params.get('search')
#         if department_code:
#             queryset = queryset.filter(department__code=department_code)
#         if status_filter:
#             queryset = queryset.filter(status=status_filter)
#         if search:
#             queryset = queryset.filter(
#                 Q(title__icontains=search) | Q(
#                     external_system_id__icontains=search)
#             )
#         return queryset.order_by('-created_at')  # Added default ordering


class BudgetProposalSummaryView(generics.GenericAPIView):
    permission_classes = [IsBMSUser]  # Allows Dept Heads
    serializer_class = BudgetProposalSummarySerializer

    @extend_schema(
        tags=['Budget Proposal Page'],
        summary="Get summary of budget proposals (cards)",
        responses={200: BudgetProposalSummarySerializer}
    )
    def get(self, request):
        user = request.user
        bms_role = get_user_bms_role(user)

        # --- NEW CODE: Filter by Current Fiscal Year ---
        today = timezone.now().date()
        current_fiscal_year = FiscalYear.objects.filter(
            start_date__lte=today,
            end_date__gte=today,
            is_active=True
        ).first()

        active_proposals = BudgetProposal.objects.filter(is_deleted=False)

        if current_fiscal_year:
            active_proposals = active_proposals.filter(
                fiscal_year=current_fiscal_year)
        # -----------------------------------------------

        # DATA ISOLATION
        if bms_role in ['ADMIN', 'FINANCE_HEAD']:
            pass  # See all
        else:
            # Dept Head sees own dept
            department_id = getattr(user, 'department_id', None)
            if department_id:
                active_proposals = active_proposals.filter(
                    department_id=department_id)
            else:
                active_proposals = BudgetProposal.objects.none()

        total = active_proposals.count()
        pending = active_proposals.filter(status='SUBMITTED').count()

        total_budget = active_proposals.aggregate(
            total=Sum('items__estimated_cost'))['total'] or 0

        data = {'total_proposals': total,
                'pending_approvals': pending, 'total_budget': total_budget}
        serializer = BudgetProposalSummarySerializer(data)
        return Response(serializer.data)


# class BudgetProposalDetailView(generics.RetrieveAPIView):
#     queryset = BudgetProposal.objects.filter(is_deleted=False).prefetch_related('items__account', 'comments')
#     serializer_class = BudgetProposalDetailSerializer
#     permission_classes = [IsBMSUser] # JWT permission

#     @extend_schema(
#         tags=['Budget Proposal Page'],
#         summary="Retrieve full details of a proposal",
#         description="Returns details including items and estimated costs.",
#         responses={200: BudgetProposalDetailSerializer}
#     )
#     def get(self, request, *args, **kwargs):
#         return super().get(request, *args, **kwargs)


# --- Proposal History Page View ---
@extend_schema(
    tags=['Proposal History Page'],
    # Changed from "List proposals history"
    summary="List proposal history entries",
    parameters=[
        OpenApiParameter(name="search", type=str), OpenApiParameter(
            name="department", type=int),
        # This status refers to ProposalHistory.action or Proposal.status?
        OpenApiParameter(name="status", type=str),
        # Assuming it filters ProposalHistory by its 'action' field.
    ],
    responses={200: ProposalHistorySerializer(many=True)}
)
# This view lists ProposalHistory entries
class ProposalHistoryView(generics.ListAPIView):
    serializer_class = ProposalHistorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = SixResultsSetPagination

    def get_queryset(self):
        qs = ProposalHistory.objects.select_related(
            'proposal__department'
        ).prefetch_related(
            'proposal__items__account__account_type'  # Add prefetch for better performance
        ).all()

        # --- MODIFICATION START: Data Isolation ---
        user = self.request.user
        bms_role = get_user_bms_role(user)

        if bms_role == 'GENERAL_USER':
            department_id = getattr(user, 'department_id', None)
            if department_id:
                # Filter history items where the parent proposal belongs to the user's department
                qs = qs.filter(proposal__department_id=department_id)
            else:
                # Fallback: If General User has no department, they see nothing
                return ProposalHistory.objects.none()
        # --- MODIFICATION END ---

        # Search by ticket ID or proposal title
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(proposal__title__icontains=search) |
                Q(proposal__external_system_id__icontains=search)
            )

        # Filter by department
        department_id = self.request.query_params.get('department')
        if department_id:
            qs = qs.filter(proposal__department_id=department_id)

        # Filter by action/status (APPROVED, REJECTED, SUBMITTED, etc.)
        action_filter = self.request.query_params.get(
            'action')  # Changed from 'status' to 'action'
        if action_filter:
            qs = qs.filter(action__iexact=action_filter)

        # Filter by category (AccountType: Asset, Expense, Liability, etc.)
        category = self.request.query_params.get('category')
        if category and category != "All Categories":
            # Filter proposals that have items with accounts of this type
            qs = qs.filter(
                proposal__items__account__account_type__name__iexact=category
            ).distinct()

        return qs.order_by('-action_at')

# --- Account Setup Page View ---


@extend_schema(
    tags=['Account Setup Page'],
    summary="List accounts with accomplishment status",
    parameters=[
        OpenApiParameter(name="fiscal_year_id", type=int, required=True),
        OpenApiParameter(name="search", type=str), OpenApiParameter(
            name="type", type=str),
        OpenApiParameter(name="status", type=str),
    ],
    responses={200: AccountSetupSerializer(many=True)}
)
# Logic seems fine, context for serializer is key
class AccountSetupListView(generics.ListAPIView):
    serializer_class = AccountSetupSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]
    # queryset defined in get_queryset to ensure it's dynamic if needed

    def get_queryset(self):
        qs = Account.objects.select_related(
            'account_type').prefetch_related('allocations')
        search = self.request.query_params.get('search')
        # Changed from 'type' to avoid clash with built-in
        acc_type_name = self.request.query_params.get('type')
        status_filter = self.request.query_params.get('status')
        if search:
            qs = qs.filter(Q(code__icontains=search) |
                           Q(name__icontains=search))
        if acc_type_name:
            qs = qs.filter(account_type__name__iexact=acc_type_name)
        if status_filter:
            if status_filter.lower() == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter.lower() == 'inactive':
                qs = qs.filter(is_active=False)
        return qs.order_by('code')

    def list(self, request, *args, **kwargs):  # Logic for passing fiscal_year context is fine
        fiscal_year_id = request.query_params.get('fiscal_year_id')
        if not fiscal_year_id:
            return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
        except FiscalYear.DoesNotExist:
            return Response({"error": "Fiscal year not found"}, status=status.HTTP_404_NOT_FOUND)
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        # Pass request to context
        serializer_context = {'request': request, 'fiscal_year': fiscal_year}
        serializer = self.get_serializer(page, many=True, context=serializer_context) if page is not None else self.get_serializer(
            queryset, many=True, context=serializer_context)
        return self.get_paginated_response(serializer.data) if page is not None else Response(serializer.data)


@extend_schema(
    tags=['Fiscal Year Dropdown'],
    summary="For Fiscal Year values in a dropdown"
)
class FiscalYearDropdownView(generics.ListAPIView):
    queryset = FiscalYear.objects.all().order_by('-start_date')
    serializer_class = FiscalYearSerializer
    permission_classes = [IsAuthenticated]


# MODIFICATION: Updated LedgerViewList to only show expense lines with proper categories

@extend_schema(
    tags=['Ledger View'],
    summary="Get ledger view",
    parameters=[
            OpenApiParameter(
                name="search", description="Search by description or category", required=False, type=str),
            OpenApiParameter(
                name="category", description="Filter by category (CAPEX or OPEX)", required=False, type=str),
            OpenApiParameter(
                name="transaction_type", description="Filter by transaction type", required=False, type=str),
            OpenApiParameter(
                name="department", description="Filter by department ID", required=False, type=int),
    ],
    responses={200: LedgerViewSerializer(many=True)}
)
class LedgerViewList(generics.ListAPIView):
    serializer_class = LedgerViewSerializer
    pagination_class = FiveResultsSetPagination
    permission_classes = [IsBMSUser]  # Changed from IsAuthenticated

    def get_queryset(self):
        queryset = JournalEntryLine.objects.select_related(
            'journal_entry',
            'account',
            'journal_entry__department',
            'expense_category'
        )

        queryset = queryset.filter(expense_category__isnull=False)

        # --- DATA ISOLATION LOGIC ---
        user = self.request.user
        bms_role = get_user_bms_role(user)

        if bms_role == 'GENERAL_USER':
            department_id = getattr(user, 'department_id', None)
            if department_id:
                # Filter lines where the parent Journal Entry belongs to the user's department
                queryset = queryset.filter(
                    journal_entry__department_id=department_id)
            else:
                return JournalEntryLine.objects.none()

        search = self.request.query_params.get('search')
        category = self.request.query_params.get('category')
        transaction_type = self.request.query_params.get('transaction_type')
        department_filter = self.request.query_params.get(
            'department_id') or self.request.query_params.get('department')

        if search:
            queryset = queryset.filter(
                Q(journal_entry__entry_id__icontains=search) |
                Q(journal_entry__date__icontains=search) |
                Q(journal_entry__description__icontains=search) |
                Q(description__icontains=search) |
                Q(expense_category__name__icontains=search) |
                Q(account__name__icontains=search) |
                Q(account__code__icontains=search)
            )

        if category:
            if category.upper() in ['CAPEX', 'OPEX']:
                queryset = queryset.filter(
                    expense_category__classification__iexact=category
                )
            else:
                queryset = queryset.filter(
                    expense_category__name__icontains=category
                )

        if transaction_type:
            queryset = queryset.filter(
                journal_transaction_type__iexact=transaction_type)

        if department_filter:
            queryset = queryset.filter(
                journal_entry__department_id=department_filter)

        return queryset.order_by('-journal_entry__date', 'journal_entry__entry_id')


@extend_schema(
    tags=['Ledger View'],
    summary="Export ledger entries as CSV",
    description="Exports filtered ledger entries in CSV format. Uses same filters as the ledger list view.",
    parameters=[
        OpenApiParameter(
            name="search", description="Search by description or category", required=False, type=str),
        OpenApiParameter(
            name="category", description="Filter by category (e.g., EXPENSES)", required=False, type=str),
        OpenApiParameter(name="transaction_type",
                         description="Filter by transaction type", required=False, type=str),
    ],
    responses={
        200: OpenApiResponse(
            description='CSV file attachment',
            response=OpenApiTypes.BINARY
        )
    }
)
class LedgerExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # This logic reuses the get_queryset from LedgerViewList for consistency
        list_view = LedgerViewList()
        list_view.request = request  # Mock the request for the view
        queryset = list_view.get_queryset()

        response = HttpResponse(
            content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="ledger_export.csv"'

        writer = csv.writer(response)
        # UPDATED: Match the new table columns
        writer.writerow(['Reference ID', 'Date', 'Category',
                        'Description', 'Account', 'Amount (PHP)'])

        for line in queryset:
            writer.writerow([
                line.journal_entry.entry_id,
                line.journal_entry.date.strftime('%Y-%m-%d'),
                line.journal_entry.category,
                line.description.replace('\n', ' ').strip(),
                line.account.name,
                "{:,.2f}".format(line.amount)
            ])
        return response

# Views of the Journal Entry page


@extend_schema(
    tags=['Journal Entry Page', 'Budget Adjustment Page'],
    summary="List Journal Entries (used for Budget Adjustment page)",
    parameters=[
            OpenApiParameter(
                "search", str, description="Search by description, Entry ID, or Account Name"),
            OpenApiParameter(
                "category", str, description="Filter by category (e.g., EXPENSES, ASSETS)")
    ],
    responses={200: JournalEntryListSerializer(many=True)}
)
class JournalEntryListView(generics.ListAPIView):
    serializer_class = JournalEntryListSerializer
    pagination_class = FiveResultsSetPagination
    # CHANGED: From IsAuthenticated to IsBMSUser
    permission_classes = [IsBMSUser]
    search_fields = ['entry_id', 'description', 'lines__account__name']

    def get_queryset(self):
        qs = JournalEntry.objects.prefetch_related(
            'lines__account', 'lines__expense_category', 'department'
        ).all()

        # --- MODIFICATION START: Data Isolation ---
        user = self.request.user
        bms_role = get_user_bms_role(user)

        if bms_role == 'GENERAL_USER':
            department_id = getattr(user, 'department_id', None)
            if department_id:
                qs = qs.filter(department_id=department_id)
            else:
                return JournalEntry.objects.none()

        search = self.request.query_params.get('search')
        category = self.request.query_params.get('category')
        department = self.request.query_params.get('department')

        if search:
            qs = qs.filter(Q(description__icontains=search) |
                           Q(entry_id__icontains=search) |
                           Q(lines__account__name__icontains=search)).distinct()

        if category:
            if category.upper() in ['CAPEX', 'OPEX']:
                qs = qs.filter(
                    lines__expense_category__classification__iexact=category).distinct()
            else:
                qs = qs.filter(category__iexact=category)

        if department:
            # Fix: Check both Name and Code
            qs = qs.filter(Q(department__name__iexact=department)
                           | Q(department__code__iexact=department))

        return qs.order_by('-date', '-entry_id')

    def get(self, *args, **kwargs):
        return super().get(*args, **kwargs)


@extend_schema(
    tags=['Journal Entry Page'],
    summary="Create a Journal Entry. Make sure to also have two lines (DEBIT,CREDIT)",
    request=JournalEntryCreateSerializer,
    responses={201: OpenApiResponse(
        description="Journal entry created successfully")}
)
# View for Journal Entry Creating
# Serializer handles user context
class JournalEntryCreateView(generics.CreateAPIView):
    serializer_class = JournalEntryCreateSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


@extend_schema(
    tags=['Account Dropdown'],
    summary="List Active Accounts",
    description="Returns a list of active accounts for use in journal entry forms or dropdowns.",
    responses={200: OpenApiResponse(
        response=AccountDropdownSerializer(many=True))}
)
class AccountDropdownView(generics.ListAPIView):
    queryset = Account.objects.filter(
        is_active=True).select_related('account_type')
    serializer_class = AccountDropdownSerializer
    # MODIFICATION START: Enable API Key access for external systems
    authentication_classes = [JWTCookieAuthentication, APIKeyAuthentication]
    permission_classes = [IsAuthenticated]
    # MODIFICATION END

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


@extend_schema(
    tags=['Journal/Ledger Category Dropdown'],
    methods=['GET'],
    summary="Get Journal Dropdown Choices",
    description="Returns valid options for journal entry categories, transaction types, and journal transaction types.",
    responses={
        200: OpenApiResponse(
            response={
                "categories": ["EXPENSES", "ASSETS", "PROJECTS", "VENDOR_CONTRACTS"],
                "transaction_types": ["DEBIT", "CREDIT"],
                "journal_transaction_types": ["CAPITAL_EXPENDITURE", "OPERATIONAL_EXPENDITURE", "TRANSFER"]
            }
        )
    }
)
@api_view(['GET'])  # journal_choices function - no change
@permission_classes([IsAuthenticated])
def journal_choices(request):
    # ... (same as your provided code)
    categories = [c[0]
                  for c in JournalEntry._meta.get_field('category').choices]
    transaction_types = [c[0] for c in JournalEntryLine._meta.get_field(
        'transaction_type').choices]
    journal_types = [c[0] for c in JournalEntryLine._meta.get_field(
        'journal_transaction_type').choices]
    return Response({"categories": categories, "transaction_types": transaction_types, "journal_transaction_types": journal_types})


@extend_schema(
    tags=['Department Dropdowns'],
    summary="List Departments",
    description="Returns a list of departments for use in filters or dropdowns.",
    responses={200: DepartmentDropdownSerializer(many=True)}
)
class DepartmentDropdownView(generics.ListAPIView):
    queryset = Department.objects.filter(is_active=True)
    serializer_class = DepartmentDropdownSerializer
    # MODIFICATION Start Enable API Key access for external systems
    authentication_classes = [JWTCookieAuthentication, APIKeyAuthentication]
    permission_classes = [IsAuthenticated]
    # MODIFICATION END

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


@extend_schema(
    tags=['Account Dropdowns'],
    summary="List Account Types",
    description="Returns all account types for filtering or dropdowns.",
    responses={200: AccountTypeDropdownSerializer(many=True)}
)
class AccountTypeDropdownView(generics.ListAPIView):
    queryset = AccountType.objects.all()
    serializer_class = AccountTypeDropdownSerializer
    # MODIFICATION START: Enable API Key access for external systems
    authentication_classes = [JWTCookieAuthentication, APIKeyAuthentication]
    permission_classes = [IsAuthenticated]
    # MODIFICATION END

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class BudgetProposalUIViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for the User Interface (UI).
    Handles listing, retrieving, reviewing, and commenting on proposals.
    """
    # Base permission: Any BMS user can read/list (subject to queryset isolation)
    permission_classes = [IsBMSUser]
    pagination_class = FiveResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = [
        'title',
        'external_system_id',
        'submitted_by_name',
        'items__cost_element',  # Sub-category (cost element)
        'items__category__name'  # Sub-category (category name)
    ]
    filterset_fields = ['status']

    def get_queryset(self):
        """
        Dynamically filters the queryset based on user role for data isolation.
        """
        user = self.request.user
        # Base query
        base_queryset = BudgetProposal.objects.filter(is_deleted=False).select_related(
            'department', 'fiscal_year'
        ).prefetch_related('items__account__account_type', 'comments', 'items__category')

        bms_role = get_user_bms_role(user)

        # 1. Determine Visibility (Data Isolation)
        if bms_role in ['ADMIN', 'FINANCE_HEAD']:
            # Admins and Finance Heads see ALL proposals
            queryset = base_queryset
        else:
            # Department Heads see ONLY their own department's proposals
            department_id = getattr(user, 'department_id', None)
            if department_id:
                queryset = base_queryset.filter(department_id=department_id)
            else:
                return BudgetProposal.objects.none()

        # 2. Apply UI Filters (Status, Category, etc.)
        status_filter = self.request.query_params.get('status')
        department_filter = self.request.query_params.get('department')
        category_filter = self.request.query_params.get('category')

        if department_filter:
            queryset = queryset.filter(department__code=department_filter)

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if category_filter:
            queryset = queryset.filter(
                items__category__classification__iexact=category_filter).distinct()

        return queryset.order_by('-created_at')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BudgetProposalDetailSerializer
        # Use MessageSerializer for creation/submission to handle nested items and logic
        if self.action == 'submit_proposal':
            return BudgetProposalMessageSerializer
        return BudgetProposalListSerializer

    # --- ACTION: SUBMIT PROPOSAL (For Dept Heads) ---
    @extend_schema(
        summary="Submit a new proposal (Department Head)",
        request=BudgetProposalMessageSerializer,
        responses={201: BudgetProposalMessageSerializer},
        tags=['Budget Proposal Page Actions']
    )
    @action(detail=False, methods=['post'], permission_classes=[CanSubmitForApproval])
    def submit_proposal(self, request):
        """
        Allows Department Heads (and others) to submit a new budget proposal.
        The proposal is created with status 'SUBMITTED' (or 'DRAFT').
        """
        # Inject the user's name as 'submitted_by_name' if not provided
        data = request.data.copy()
        if 'submitted_by_name' not in data:
            data['submitted_by_name'] = request.user.get_full_name(
            ) or request.user.username

        # If user is a Dept Head, force the department to be their own
        bms_role = get_user_bms_role(request.user)
        if bms_role == 'GENERAL_USER':
            department_id = getattr(request.user, 'department_id', None)
            if department_id:
                data['department_input'] = str(department_id)

        serializer = BudgetProposalMessageSerializer(
            data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        proposal = serializer.save()

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    # --- ACTION: REVIEW (Finance Head Only) ---
    @extend_schema(
        summary="Review a proposal (Finance Head)",
        request=ProposalReviewSerializer,
        responses={200: BudgetProposalDetailSerializer},
        tags=['Budget Proposal Page Actions']
    )
    @action(detail=True, methods=['post'], permission_classes=[IsBMSFinanceHead])
    def review(self, request, pk=None):
        proposal = self.get_object()

        review_input_serializer = ProposalReviewSerializer(data=request.data)
        review_input_serializer.is_valid(raise_exception=True)

        new_status = review_input_serializer.validated_data['status']
        comment_text = review_input_serializer.validated_data.get('comment', '')
        finance_operator = review_input_serializer.validated_data.get('finance_manager_name', '')
        signature_file = review_input_serializer.validated_data.get('signature')

        reviewer_user_id = request.user.id
        reviewer_name = request.user.get_full_name() or request.user.username
        previous_status_for_history = proposal.status

        with transaction.atomic():
            proposal.status = new_status

            if finance_operator:
                proposal.finance_manager_name = finance_operator
            if signature_file:
                proposal.signature = signature_file

            if new_status == 'APPROVED':
                proposal.approved_by_name = reviewer_name
                proposal.approval_date = timezone.now()
                proposal.rejected_by_name = None
                proposal.rejection_date = None

                # ✅ FIX 1: Get or Create Project (prevents duplicates)
                if not hasattr(proposal, 'project') or proposal.project is None:
                    project = Project.objects.create(
                        name=f"Project for: {proposal.title}",
                        description=proposal.project_summary,
                        start_date=proposal.performance_start_date,
                        end_date=proposal.performance_end_date,
                        department=proposal.department,
                        budget_proposal=proposal,
                        status='PLANNING'
                    )
                else:
                    project = proposal.project

                # ✅ FIX 2: Create BudgetAllocations from Proposal Items
                # This is the CRITICAL missing piece - without allocations, projects have $0 budget
                for item in proposal.items.all():
                    # Validate that category exists (should be caught by serializer, but double-check)
                    if not item.category:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Item {item.id} has no category. Skipping allocation creation.")
                        continue

                    # Check if allocation already exists (prevents duplicates on re-approval)
                    existing_allocation = BudgetAllocation.objects.filter(
                        project=project,
                        category=item.category,
                        account=item.account,
                        is_active=True
                    ).first()

                    if not existing_allocation:
                        BudgetAllocation.objects.create(
                            fiscal_year=proposal.fiscal_year,
                            department=proposal.department,
                            category=item.category,
                            account=item.account,
                            project=project,
                            proposal=proposal,
                            amount=item.estimated_cost,
                            created_by_name=reviewer_name,
                            is_active=True,
                            is_locked=False  # ✅ Unlocked = Ready for expenses
                        )
                    else:
                        # If re-approving, update amount and unlock
                        existing_allocation.amount = item.estimated_cost
                        existing_allocation.is_locked = False
                        existing_allocation.save(update_fields=['amount', 'is_locked'])

            elif new_status == 'REJECTED':
                proposal.rejected_by_name = reviewer_name
                proposal.rejection_date = timezone.now()
                proposal.approved_by_name = None
                proposal.approval_date = None
                
                if hasattr(proposal, 'project') and proposal.project is not None:
                    proposal.project.status = 'CANCELLED'
                    proposal.project.save(update_fields=['status'])

            proposal.last_modified = timezone.now()
            proposal.save()

            if comment_text:
                ProposalComment.objects.create(
                    proposal=proposal,
                    user_id=reviewer_user_id,
                    user_username=reviewer_name,
                    comment=comment_text
                )

            ProposalHistory.objects.create(
                proposal=proposal, action=new_status,
                action_by_name=reviewer_name,
                previous_status=previous_status_for_history,
                new_status=proposal.status,
                comments=comment_text or f"Status changed to {new_status}."
            )

        # --- External System Notification (Existing Code) ---
        if proposal.external_system_id:
            try:
                target_url = None
                api_key_to_use = None
                
                tid = proposal.external_system_id.upper()
                
                if tid.startswith("AST") or tid.startswith("REP") or tid.startswith("REG"):
                     target_url = getattr(settings, 'AMS_STATUS_UPDATE_URL', None)
                     api_key_to_use = getattr(settings, 'API_KEY_FOR_BMS_TO_CALL_AMS', None)
                
                elif tid.startswith("HD") or tid.startswith("TRV"):
                     target_url = getattr(settings, 'HDTS_STATUS_UPDATE_URL', None)
                     api_key_to_use = getattr(settings, 'API_KEY_FOR_BMS_TO_CALL_HDTS', None)

                else:
                     target_url = getattr(settings, 'DTS_STATUS_UPDATE_URL', None)
                     api_key_to_use = getattr(settings, 'BMS_AUTH_KEY_FOR_DTS', None)

                if target_url:
                    payload = {
                        'ticket_id': proposal.external_system_id,
                        'status': new_status,
                        'comment': comment_text,
                        'reviewed_by': reviewer_name,
                        'reviewed_at': timezone.now().isoformat(),
                        'order_number': proposal.external_system_id 
                    }
                    headers = {
                        'Content-Type': 'application/json',
                        'X-API-Key': api_key_to_use or ''
                    }
                    
                    response = requests.post(target_url, json=payload, headers=headers, timeout=5)
                    
                    if 200 <= response.status_code < 300:
                        proposal.sync_status = 'SYNCED'
                    else:
                        print(f"External system returned {response.status_code}: {response.text}")
                        proposal.sync_status = 'FAILED'
                        
                    proposal.last_sync_timestamp = timezone.now()
                    proposal.save(update_fields=['sync_status', 'last_sync_timestamp'])
                else:
                    print(f"Warning: No callback URL configured for prefix {tid}. Skipping notification.")

            except requests.RequestException as e:
                print(f"Error notifying external system for proposal {proposal.id}: {e}")
                proposal.sync_status = 'FAILED'
                proposal.save(update_fields=['sync_status'])

        output_serializer = BudgetProposalDetailSerializer(
            proposal, context={'request': request})
        return Response(output_serializer.data, status=status.HTTP_200_OK)

    # --- ACTION: ADD COMMENT (All Users) ---
    @extend_schema(
        summary="Add a comment to a proposal",
        request=ProposalCommentCreateSerializer,
        responses={201: ProposalCommentSerializer},
        tags=['Budget Proposal Page Actions']
    )
    @action(detail=True, methods=['post'], url_path='add-comment', permission_classes=[IsBMSUser])
    def add_comment(self, request, pk=None):
        proposal = self.get_object()
        comment_serializer = ProposalCommentCreateSerializer(data=request.data)
        comment_serializer.is_valid(raise_exception=True)
        comment_text = comment_serializer.validated_data['comment']

        commenter_name = f"{request.user.first_name} {request.user.last_name}".strip(
        ) or request.user.username

        comment_obj = ProposalComment.objects.create(
            proposal=proposal,
            user_id=request.user.id,
            user_username=commenter_name,
            comment=comment_text
        )
        ProposalHistory.objects.create(
            proposal=proposal, action='COMMENTED', action_by_name=commenter_name,
            comments=f"Comment added: '{comment_text}'"
        )
        return Response(ProposalCommentSerializer(comment_obj).data, status=status.HTTP_201_CREATED)

@extend_schema(tags=['External System Integration (API Key Protected)'])
class ExternalBudgetProposalViewSet(mixins.CreateModelMixin,
                                   mixins.UpdateModelMixin,
                                   mixins.RetrieveModelMixin,
                                   viewsets.GenericViewSet):
    """
    ViewSet for external systems (DTS/TTS/AMS/HDTS) to create and manage budget proposals.
    
    **Authentication**: Requires valid API Key in `X-API-Key` header.
    
    **Endpoints**:
    - POST /external/budget-proposals/ - Create a new proposal
    - PUT /external/budget-proposals/{external_system_id}/ - Update existing proposal
    - PATCH /external/budget-proposals/{external_system_id}/ - Partial update
    - GET /external/budget-proposals/{external_system_id}/ - Retrieve proposal status
    - POST /external/budget-proposals/{external_system_id}/cancel/ - Cancel a proposal
    """
    queryset = BudgetProposal.objects.filter(is_deleted=False)
    serializer_class = BudgetProposalMessageSerializer
    authentication_classes = [APIKeyAuthentication]
    permission_classes = [IsTrustedService]
    lookup_field = 'external_system_id'
    
    def get_service_name(self):
        """Extract service name from the authenticated service principal"""
        if hasattr(self.request.user, 'service_name'):
            return self.request.user.service_name
        return "External System"

    @extend_schema(
        summary="Create a new budget proposal",
        description="""
        Create a new budget proposal from an external system.
        
        **Required Fields**:
        - `ticket_id`: Unique ID from your system (will be used as `external_system_id`)
        - `title`: Proposal title
        - `department_input`: Department ID or Code
        - `fiscal_year`: Fiscal Year ID
        - `items`: Array of budget items
        
        **Optional Fields**:
        - `is_draft`: Set to `true` to save as draft (default: false = submitted)
        - `submitted_by_name`: Name of the requester
        - `document`: Attached supporting document
        
        **Response**: Returns the created proposal with BMS-assigned ID.
        """,
        request=BudgetProposalMessageSerializer,
        responses={
            201: OpenApiResponse(
                response=BudgetProposalMessageSerializer,
                description="Proposal created successfully"
            ),
            400: OpenApiResponse(description="Validation error"),
            403: OpenApiResponse(description="Invalid API Key")
        },
        examples=[
            OpenApiExample(
                "Create Proposal Example",
                value={
                    "ticket_id": "HD-2026-001",
                    "title": "Server Infrastructure Upgrade",
                    "project_summary": "Upgrade data center servers",
                    "project_description": "Replace aging servers with new hardware",
                    "department_input": "IT",
                    "fiscal_year": 1,
                    "submitted_by_name": "John Doe",
                    "performance_start_date": "2026-02-01",
                    "performance_end_date": "2026-06-30",
                    "items": [
                        {
                            "cost_element": "Server Hardware",
                            "description": "Dell PowerEdge R750",
                            "estimated_cost": "150000.00",
                            "account": 5,
                            "notes": "Qty: 3 units"
                        }
                    ],
                    "is_draft": False
                }
            )
        ]
    )
    def create(self, request, *args, **kwargs):
        """Create a new budget proposal"""
        service_name = self.get_service_name()
        
        # Log the incoming request for debugging
        print(f"📥 Incoming proposal from {service_name}: {request.data.get('ticket_id')}")
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Save the proposal
        proposal = serializer.save()
        
        # Return created proposal
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "message": "Budget proposal created successfully",
                "proposal_id": proposal.id,
                "external_system_id": proposal.external_system_id,
                "status": proposal.status,
                "data": serializer.data
            },
            status=status.HTTP_201_CREATED,
            headers=headers
        )

    @extend_schema(
        summary="Update an existing budget proposal",
        description="""
        Update a proposal that's still in DRAFT or SUBMITTED status.
        
        **Note**: Approved or Rejected proposals cannot be updated.
        Use the `external_system_id` (your ticket ID) to identify the proposal.
        """,
        request=BudgetProposalMessageSerializer,
        responses={
            200: OpenApiResponse(description="Proposal updated successfully"),
            400: OpenApiResponse(description="Validation error or proposal is locked"),
            404: OpenApiResponse(description="Proposal not found")
        }
    )
    def update(self, request, *args, **kwargs):
        """Update an existing proposal"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Prevent updates to finalized proposals
        if instance.status in ['APPROVED', 'REJECTED']:
            return Response(
                {"error": f"Cannot update proposal with status '{instance.status}'"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        proposal = serializer.save()
        
        return Response({
            "message": "Budget proposal updated successfully",
            "proposal_id": proposal.id,
            "external_system_id": proposal.external_system_id,
            "status": proposal.status,
            "data": serializer.data
        })

    @extend_schema(
        summary="Retrieve proposal status",
        description="""
        Get the current status and details of a proposal.
        
        **Use this endpoint to**:
        - Check if a proposal has been approved/rejected
        - Get feedback comments from Finance Manager
        - Retrieve the BMS-assigned proposal ID
        """,
        responses={
            200: BudgetProposalMessageSerializer,
            404: OpenApiResponse(description="Proposal not found")
        }
    )
    def retrieve(self, request, *args, **kwargs):
        """Retrieve proposal details"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        
        # Add useful metadata
        return Response({
            "proposal_id": instance.id,
            "external_system_id": instance.external_system_id,
            "status": instance.status,
            "submitted_at": instance.submitted_at,
            "last_modified": instance.last_modified,
            "approved_by": instance.approved_by_name,
            "approval_date": instance.approval_date,
            "rejected_by": instance.rejected_by_name,
            "rejection_date": instance.rejection_date,
            "sync_status": instance.sync_status,
            "data": serializer.data
        })

    @extend_schema(
        summary="Cancel a budget proposal",
        description="""
        Mark a proposal as deleted (soft delete).
        
        **Only allowed for**:
        - DRAFT proposals
        - SUBMITTED proposals (before review)
        
        **Not allowed for**:
        - APPROVED proposals (projects may already be created)
        - REJECTED proposals (already finalized)
        """,
        request=None,
        responses={
            200: OpenApiResponse(description="Proposal cancelled successfully"),
            400: OpenApiResponse(description="Cannot cancel this proposal"),
            404: OpenApiResponse(description="Proposal not found")
        }
    )
    @action(detail=True, methods=['post'])
    def cancel(self, request, external_system_id=None):
        """Cancel a proposal (soft delete)"""
        proposal = self.get_object()
        service_name = self.get_service_name()
        
        # Only allow cancellation of pending proposals
        if proposal.status in ['APPROVED', 'REJECTED']:
            return Response(
                {"error": f"Cannot cancel a {proposal.status.lower()} proposal"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            proposal.is_deleted = True
            proposal.save()
            
            # Log the cancellation
            ProposalHistory.objects.create(
                proposal=proposal,
                action='UPDATED',
                action_by_name=f"System ({service_name})",
                new_status='CANCELLED',
                previous_status=proposal.status,
                comments=f"Proposal cancelled by {service_name}"
            )
        
        return Response({
            "message": "Proposal cancelled successfully",
            "proposal_id": proposal.id,
            "external_system_id": proposal.external_system_id
        })

    @extend_schema(
        summary="Check proposal status (quick check)",
        description="""
        Lightweight endpoint to quickly check if a proposal is approved.
        
        **Returns**:
        - `exists`: Boolean - Does the proposal exist?
        - `status`: Current status (DRAFT/SUBMITTED/APPROVED/REJECTED)
        - `is_approved`: Boolean - Quick check for approval
        - `order_number`: The external_system_id you can use for expense tracking
        """,
        responses={
            200: OpenApiResponse(
                description="Status information",
                examples=[
                    OpenApiExample(
                        "Approved Proposal",
                        value={
                            "exists": True,
                            "status": "APPROVED",
                            "is_approved": True,
                            "order_number": "HD-2026-001",
                            "approved_at": "2026-01-15T10:30:00Z"
                        }
                    )
                ]
            ),
            404: OpenApiResponse(description="Proposal not found")
        }
    )
    @action(detail=True, methods=['get'], url_path='status-check')
    def status_check(self, request, external_system_id=None):
        """Quick status check for external systems"""
        try:
            proposal = self.get_object()
            
            return Response({
                "exists": True,
                "status": proposal.status,
                "is_approved": proposal.status == 'APPROVED',
                "is_rejected": proposal.status == 'REJECTED',
                "order_number": proposal.external_system_id,
                "submitted_at": proposal.submitted_at,
                "approved_at": proposal.approval_date,
                "rejected_at": proposal.rejection_date,
                "approved_by": proposal.approved_by_name,
                "rejected_by": proposal.rejected_by_name
            })
        except BudgetProposal.DoesNotExist:
            return Response(
                {
                    "exists": False,
                    "message": f"No proposal found with ID: {external_system_id}"
                },
                status=status.HTTP_404_NOT_FOUND
            )


@extend_schema(
    tags=["Budget Proposal Export"],
    summary="Export a budget proposal to Excel",
    description="Exports the specified BudgetProposal and its items to an Excel file. Example:",
    responses={
        200: OpenApiResponse(description="XLSX file of the proposal"),
        404: OpenApiResponse(description="Proposal not found"),
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_budget_proposal_excel(request, proposal_id):
    try:
        proposal = BudgetProposal.objects.prefetch_related('items__account', 'department').get(
            id=proposal_id, is_deleted=False)  # Added is_deleted=False
    except BudgetProposal.DoesNotExist:
        return HttpResponse("Proposal not found", status=404)

    # --- START: Optional User Activity Logging ---
    try:
        UserActivityLog.objects.create(
            user_id=request.user.id,  # From JWT
            user_username=getattr(request.user, 'username', 'N/A'),  # From JWT
            log_type='EXPORT',  # Ensure 'EXPORT' is in UserActivityLog.LOG_TYPE_CHOICES
            action=f'Exported budget proposal to Excel: ID {proposal.id}, Title "{proposal.title}"',
            status='SUCCESS',
            details={
                'proposal_id': proposal.id,
                'proposal_title': proposal.title,
                'export_format': 'xlsx'
            }
        )
    except Exception as e:
        # Log an error if activity logging fails, but don't let it break the export
        print(f"Error logging export activity: {e}")
    # --- END: Optional User Activity Logging ---

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Proposal"

    # Header section
    header_labels = [
        "Title", "Project Summary", "Project Description",
        "Performance Start", "Performance End", "Performance Notes",
        "Department", "Submitted By", "Status"
    ]
    header_values = [
        proposal.title,
        proposal.project_summary,
        proposal.project_description,
        proposal.performance_start_date.strftime("%Y-%m-%d"),
        proposal.performance_end_date.strftime("%Y-%m-%d"),
        getattr(proposal, 'performance_notes', ''),
        proposal.department.name,
        proposal.submitted_by_name or "N/A",
        proposal.status
    ]
    for i, (label, value) in enumerate(zip(header_labels, header_values), start=1):
        ws.append([label, value])
        ws.cell(row=i, column=1).font = Font(
            bold=True)  # Bold the header label

    ws.append([])  # Blank row

    # able Header
    table_header = ["Account", "Cost Element",
                    "Description", "Estimated Cost", "Notes"]
    ws.append(table_header)
    table_header_row = ws.max_row
    for col in range(1, len(table_header) + 1):
        ws.cell(row=table_header_row, column=col).font = Font(bold=True)

    # Line Items
    total_cost = 0
    for item in proposal.items.all():
        ws.append([
            item.account.name if item.account else "N/A",
            item.cost_element,
            item.description,
            float(item.estimated_cost),
            item.notes or ""
        ])
        total_cost += item.estimated_cost

    # Total Row
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["", "", "Total", float(total_cost)])

    # Bold "Total", "Estimated Cost" header, and the summed value
    ws.cell(row=table_header_row, column=4).font = Font(
        bold=True)  # "Estimated Cost" header
    ws.cell(row=table_header_row, column=5).font = Font(
        bold=True)  # "Notes" header
    ws.cell(row=total_row, column=3).font = Font(
        bold=True)         # "Total" label
    ws.cell(row=total_row, column=4).font = Font(
        bold=True)         # summed value

    # Bold all "Description" column cells (header and values)
    for row in range(table_header_row, ws.max_row + 1):
        ws.cell(row=row, column=3).font = Font(bold=True)

    # Auto size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = max(max_length + 2, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"budget_proposal_{proposal.id}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


# MODIFICATION START
'''
Business Logic Explanation for BudgetVarianceReportView:

This view generates a hierarchical budget variance report, for tracking performance
It compares planned financial outflows (Budget) against actual financial outflows (Actual)

Implemantation:
-   Structure: The report's nested structure is built recursively based on the parent-child relationships 
    defined in the `ExpenseCategory` model. The process starts with top-level categories (e.g., 'INCOME', 'EXPENSE')

-   'Budget' Calculation: For each category, the `Budget` is the sum of all `BudgetAllocation` amounts linked to 
    that specific category (and its sub-categories) for the selected `fiscal_year`.  This represents the total 
    planned allocation.

-   'Actual' Calculation: The `Actual` is the sum of all `APPROVED` `Expense` amounts linked to that category
    If a `month` parameter is provided in the request, this calculation is filtered to only include expenses 
    from that specific month. This represents the total recorded spending.

-   'Available' Calculation: The `Available` amount is a simple calculation: `Budget - Actual`. A positive 
    value indicates underspending, while a negative value (shown in red on the UI) indicates overspending


Business Logic Consideration for ASSET, LIABILITY, and EXPENSE categories:

This view uses a unified calculation (`Budget - Actual`) across all top-level categories, which is a logical 
simplification for this application's purpose. Here is how to interpret the results for each type:

-   For 'EXPENSE' categories (e.g., Operations, Marketing): This is a standard budget variance calculation. It shows 
    how much of the operational budget has been spent and how much remains.

-   For 'ASSET' categories (e.g., Equipment): This represents a capital expenditure (CapEx) budget. The `Budget` is 
    the total allocated for purchasing assets, `Actual` is the value of assets already purchased against that budget, 
    and `Available` correctly shows the remaining funds for future asset acquisition.

-   For 'LIABILITY' categories (e.g., Payables and Loans): This can be interpreted as tracking the budget allocated 
    for debt servicing. The `Budget` is the amount the company has set aside to make payments on its obligations, 
    `Actual` represents the payments that have been made, and `Available` shows the remaining funds allocated for 
    this purpose within the period.

this unified approach is effective and consistent 
for this application's goal: tracking planned financial outflows against actual outflows across all defined categories
'''


class BudgetVarianceReportView(APIView):
    permission_classes = [IsBMSUser]  # Changed from IsAuthenticated
    '''
    View for budget variance report page
    '''
    @extend_schema(
        tags=["Budget Variance Reports"],
        summary="Budget Variance Report",
        description="The Budget Variance Report shows how budget allocations are utilized...",
        parameters=[
            OpenApiParameter(name="fiscal_year_id", required=True, type=int),
            OpenApiParameter(
                name="month", description="Filter actuals by a specific month (1-12).", type=int)
        ],
        responses={200: ExpenseCategoryVarianceSerializer(many=True)}
    )
    def get(self, request):
        fiscal_year_id = request.query_params.get('fiscal_year_id')
        if not fiscal_year_id:
            return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(request.query_params.get('month')
                        ) if request.query_params.get('month') else None
            if month and (month < 1 or month > 12):
                raise ValueError()
        except (ValueError, TypeError):
            return Response({"error": "Invalid month provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
        except FiscalYear.DoesNotExist:
            return Response({"error": "Fiscal Year not found"}, status=status.HTTP_404_NOT_FOUND)

        # --- DATA ISOLATION LOGIC ---
        user = request.user
        bms_role = get_user_bms_role(user)

        filter_dept_id = None
        if bms_role == 'GENERAL_USER':
            filter_dept_id = getattr(user, 'department_id', None)
            if not filter_dept_id:
                # Dept Head with no department sees empty report
                return Response([])

        try:
            UserActivityLog.objects.create(
                user_id=request.user.id,  # From JWT
                user_username=getattr(
                    request.user, 'username', 'N/A'),  # From JWT
                log_type='REPORT_VIEW',  # Or 'REPORT_GENERATE', ensure type exists
                action=f'Generated Budget Variance Report for FY: {fiscal_year.name}',
                status='SUCCESS',
                details={
                    'fiscal_year_id': fiscal_year.id,
                    'fiscal_year_name': fiscal_year.name,
                    'report_type': 'BudgetVariance'
                }
            )
        except Exception as e:
            print(f"Error logging report generation activity: {e}")
        # --- END: Optional User Activity Logging ---

        top_categories = ExpenseCategory.objects.filter(
            level=1, is_active=True)

        def aggregate_node(category):
            # 1. Budget Calculation
            budget_qs = BudgetAllocation.objects.filter(
                category=category, fiscal_year=fiscal_year, is_active=True
            )
            # Apply Dept Filter
            if filter_dept_id:
                budget_qs = budget_qs.filter(department_id=filter_dept_id)

            budget = budget_qs.aggregate(total=Coalesce(
                Sum('amount'), 0, output_field=DecimalField()))['total']

            # 2. Actual Calculation
            actual_qs = Expense.objects.filter(
                category=category,
                budget_allocation__fiscal_year=fiscal_year,
                status='APPROVED'
            )

            # Apply Dept Filter
            if filter_dept_id:
                actual_qs = actual_qs.filter(department_id=filter_dept_id)

            if month:
                actual_qs = actual_qs.filter(date__month=month)

            actual = actual_qs.aggregate(total=Coalesce(
                Sum('amount'), 0, output_field=DecimalField()))['total']

            available = budget - actual
            children = []

            for child in category.subcategories.all():
                children.append(aggregate_node(child))

            # Aggregate child totals if exists
            if children:
                child_budget = sum(c['budget'] for c in children)
                child_actual = sum(c['actual'] for c in children)
                child_available = child_budget - child_actual
                return {
                    "category": category.name,
                    "code": category.code,
                    "level": category.level,
                    "classification": category.classification,  # Ensure field exists
                    "budget": round(child_budget, 2),
                    "actual": round(child_actual, 2),
                    "available": round(child_available, 2),
                    "children": children
                }
            else:
                return {
                    "category": category.name,
                    "code": category.code,
                    "level": category.level,
                    "classification": category.classification,  # Ensure field exists
                    "budget": round(budget, 2),
                    "actual": round(actual, 2),
                    "available": round(available, 2),
                    "children": []
                }
        # Builds nested dictionariers, each representing top level expense categories, and its full budget/actual/available breakdown (and subcategories)
        # aggregate_node(cat) - recursive function that computes budget, actual, and available amounts for given category (and all its children if any)
        result = [aggregate_node(cat) for cat in top_categories]
        return Response(result)


@extend_schema(
    tags=["Budget Proposal Page Actions"],
    summary="Get budget overview for a proposal review",
    description="Provides financial context for a department's budget when reviewing a specific proposal.",
    responses={200: ProposalReviewBudgetOverviewSerializer}
)
class ProposalReviewBudgetOverview(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, proposal_id):
        try:
            proposal = BudgetProposal.objects.select_related(
                'department', 'fiscal_year').get(id=proposal_id)
        except BudgetProposal.DoesNotExist:
            return Response({"error": "Proposal not found"}, status=status.HTTP_404_NOT_FOUND)

        department = proposal.department
        fiscal_year = proposal.fiscal_year
        proposal_cost = proposal.items.aggregate(
            total=Coalesce(Sum('estimated_cost'), Decimal('0')))['total']

        # 1. Total budget allocated to the department for the fiscal year
        total_dept_budget = BudgetAllocation.objects.filter(
            department=department, fiscal_year=fiscal_year, is_active=True
        ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

        # 2. Total spent by the department so far
        total_spent = Expense.objects.filter(
            department=department, budget_allocation__fiscal_year=fiscal_year, status='APPROVED'
        ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

        # UI's "Currently Allocated" seems to mean "already spent"
        currently_allocated = total_spent

        # 3. Available budget before considering this new proposal
        available_budget = total_dept_budget - total_spent

        # 4. Budget after this proposal is approved
        budget_after_proposal = available_budget - proposal_cost

        data = {
            "total_department_budget": total_dept_budget,
            "currently_allocated": currently_allocated,
            "available_budget": available_budget,
            "budget_after_proposal": budget_after_proposal
        }

        serializer = ProposalReviewBudgetOverviewSerializer(data)
        return Response(serializer.data)


@extend_schema(
    tags=["Budget Variance Reports"],
    summary="Export Budget Variance Report to Excel",
    description="Exports the hierarchical budget variance report to an XLSX file.",
    parameters=[
        OpenApiParameter(name="fiscal_year_id", required=True, type=int)
    ],
    responses={
        200: OpenApiResponse(description="XLSX file of the budget variance report"),
        400: OpenApiResponse(description="fiscal_year_id is required"),
        404: OpenApiResponse(description="Fiscal Year not found"),
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_budget_variance_excel(request):
    fiscal_year_id = request.query_params.get('fiscal_year_id')
    if not fiscal_year_id:
        return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)

    # ADDED: Parse month parameter (same as display endpoint)
    try:
        month = int(request.query_params.get('month')
                    ) if request.query_params.get('month') else None
        if month and (month < 1 or month > 12):
            raise ValueError()
    except (ValueError, TypeError):
        return Response(
            {"error": "Invalid month provided. Must be an integer between 1 and 12."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
    except FiscalYear.DoesNotExist:
        return Response({"error": "Fiscal Year not found"}, status=status.HTTP_404_NOT_FOUND)

    # Re-use the logic from BudgetVarianceReportView to get the data
    top_categories = ExpenseCategory.objects.filter(level=1, is_active=True)

    def aggregate_node(category):
        # Budget (whole year, not filtered by month)
        budget = BudgetAllocation.objects.filter(
            category=category, fiscal_year=fiscal_year, is_active=True
        ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

        # FIXED: Apply month filter to actual expenses (same as display endpoint)
        actual_qs = Expense.objects.filter(
            category=category,
            budget_allocation__fiscal_year=fiscal_year,
            status='APPROVED'
        )

        # Apply month filter if provided
        if month:
            actual_qs = actual_qs.filter(date__month=month)

        actual = actual_qs.aggregate(
            total=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )['total']

        children_data = [aggregate_node(child)
                         for child in category.subcategories.all()]

        if children_data:
            child_budget = sum(c['budget'] for c in children_data)
            child_actual = sum(c['actual'] for c in children_data)
            return {
                "name": category.name,
                "budget": child_budget,
                "actual": child_actual,
                "available": child_budget - child_actual,
                "children": children_data
            }
        else:
            return {
                "name": category.name,
                "budget": budget,
                "actual": actual,
                "available": budget - actual,
                "children": []
            }

    report_data = [aggregate_node(cat) for cat in top_categories]

    # --- Excel Generation ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Variance Report"

    # IMPROVED: Add month info to title if filtering
    if month:
        ws.append(
            [f"Budget Variance Report - {fiscal_year.name} (Month: {month})"])
    else:
        ws.append([f"Budget Variance Report - {fiscal_year.name} (Full Year)"])
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([])  # Blank row

    header = ["Category", "Budget", "Actual", "Available"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=3, column=col).font = Font(bold=True)
        # IMPROVED: Add currency number format
        if col > 1:  # Budget, Actual, Available columns
            ws.cell(row=3, column=col).number_format = '#,##0.00'

    def write_rows(data, indent=0, start_row=4):
        current_row = start_row
        for item in data:
            ws.append([
                f"{' ' * indent * 4}{item['name']}",
                item['budget'],
                item['actual'],
                item['available']
            ])
            # Apply currency format
            for col in range(2, 5):  # Columns B, C, D
                ws.cell(row=current_row, column=col).number_format = '#,##0.00'

            current_row += 1
            if item.get('children'):
                current_row = write_rows(
                    item['children'], indent + 1, current_row)[1]
        return current_row, current_row

    write_rows(report_data)

    # Auto-size columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # IMPROVED: Include month in filename
    if month:
        filename = f"budget_variance_report_{fiscal_year.name}_month{month}.xlsx"
    else:
        filename = f"budget_variance_report_{fiscal_year.name}.xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@extend_schema(
    tags=["Budget Adjustment Page"],
    summary="Create a Budget Adjustment",
    request=BudgetAdjustmentSerializer,
    responses={201: JournalEntryListSerializer}
)
class BudgetAdjustmentView(generics.CreateAPIView):
    # Only Finance Heads can legally modify the budget allocation.
    permission_classes = [IsBMSFinanceHead]
    serializer_class = BudgetAdjustmentSerializer

    def perform_create(self, serializer):
        data = serializer.validated_data
        user = self.request.user
        amount = data['amount']

        # --- NEW CODE: Extract Transfer Type ---
        transfer_type = data.get('transfer_type', 'TRANSFER')
        # ---------------------------------------

        source_alloc = data.get('source_alloc')
        dest_alloc = data.get('dest_alloc')
        dept = data['department']

        # --- NEW CODE: Description Update ---
        description = data.get(
            'description') or f"Budget {transfer_type.title()}"
        # ------------------------------------

        with transaction.atomic():
            # 1. Update Allocations (Real Impact)

            # --- NEW CODE: Handle Transfer vs Supplemental ---
            if transfer_type == 'TRANSFER':
                if source_alloc:
                    source_alloc.amount -= amount  # Reduce source
                    if source_alloc.amount < 0:
                        raise serializers.ValidationError(
                            "Source allocation cannot go negative")
                    source_alloc.save()

            # For Supplemental, we DO NOT deduct from source (it's new money)
            # -------------------------------------------------

            if dest_alloc:
                dest_alloc.amount += amount  # Increase destination
                dest_alloc.save()

            # 2. Create Journal Entry (Audit)
            je = JournalEntry.objects.create(
                date=data['date'],
                category='PROJECTS',
                description=description,
                total_amount=amount,
                status='POSTED',
                department=dept,
                created_by_user_id=user.id,
                created_by_username=getattr(user, 'username', 'N/A')
            )

            # --- NEW CODE: Source Account Logic for JE ---
            source_account = data.get('source_account_obj')

            if transfer_type == 'SUPPLEMENTAL':
                # If Supplemental, we credit a generic "Treasury" or "Equity" account
                # because the money isn't coming from another allocation.
                if not source_account:
                    # Fallback lookup for a system account
                    source_account = Account.objects.filter(
                        Q(name__icontains='Treasury') | Q(
                            account_type__name='Equity')
                    ).first()

            if not source_account:
                # Safety fallback to prevent crash, though in production this should be configured
                source_account = Account.objects.filter(is_active=True).first()
            # ---------------------------------------------

            # For Asset Accounts (Budget Accounts):
            # CREDIT the source (money decreasing OR Equity increasing)
            JournalEntryLine.objects.create(
                journal_entry=je,
                account=source_account,
                transaction_type='CREDIT',  # Decrease in asset / Increase in Equity
                journal_transaction_type='TRANSFER',
                amount=amount,
                description=f"{transfer_type.title()} source: {source_account.name}",
                expense_category=source_alloc.category if source_alloc else None
            )

            # DEBIT the destination (money increasing)
            JournalEntryLine.objects.create(
                journal_entry=je,
                account=data['destination_account_obj'],
                transaction_type='DEBIT',  # Increase in asset
                journal_transaction_type='TRANSFER',
                amount=amount,
                description=f"{transfer_type.title()} to {data['destination_account_obj'].name}",
                expense_category=dest_alloc.category if dest_alloc else None
            )

            self.created_instance = je

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        response_serializer = JournalEntryListSerializer(self.created_instance)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

# MODIFICATION START: Endpoint for Operators to Request Supplemental Budget


@extend_schema(
    tags=["Supplemental Budget"],
    summary="Request Supplemental Budget (Operator)",
    description="Creates a pending BudgetTransfer. Automatically creates a 0-balance allocation bucket if one does not exist for the Project+Category.",
    request=SupplementalBudgetRequestSerializer,
    responses={201: OpenApiResponse(
        description="Request submitted successfully")}
)
class SupplementalBudgetRequestView(generics.CreateAPIView):
    permission_classes = [IsBMSUser]
    serializer_class = SupplementalBudgetRequestSerializer

    def perform_create(self, serializer):
        data = serializer.validated_data
        user = self.request.user

        allocation = data.get('allocation_obj')

        # If allocation doesn't exist, create a placeholder one ($0)
        if not allocation:
            with transaction.atomic():
                # Simplified Account selection: Use default Expense account
                account = Account.objects.filter(
                    account_type__name='Expense').first()
                if not account:
                    # Absolute fallback
                    account = Account.objects.first()

                allocation = BudgetAllocation.objects.create(
                    fiscal_year=data['fiscal_year_obj'],
                    department=data['department_obj'],
                    category=data['category_obj'],
                    account=account,
                    project=data['project_obj'],
                    amount=Decimal('0.00'),
                    created_by_name='System (Supplemental Request)',
                    is_active=True,
                    is_locked=False  # Unlocked so it can receive funds
                )

        # Create Transfer Record with status PENDING
        BudgetTransfer.objects.create(
            fiscal_year=data['fiscal_year_obj'],
            source_allocation=None,
            destination_allocation=allocation,
            amount=data['amount'],
            reason=data['reason'],
            transfer_type='SUPPLEMENTAL',
            status='PENDING',
            transferred_by_user_id=user.id,
            transferred_by_username=getattr(user, 'username', 'N/A')
        )

        # Optional: Log activity
        UserActivityLog.objects.create(
            user_id=user.id,
            user_username=getattr(user, 'username', 'N/A'),
            log_type='CREATE',
            action=f"Requested supplemental budget of {data['amount']} for {data['department_obj'].code}",
            status='SUCCESS',
            details={'category': data['category_obj'].name,
                     'project': data['project_obj'].name}
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({"message": "Supplemental budget request submitted for approval."}, status=status.HTTP_201_CREATED)
# MODIFICATION END


class ExternalJournalEntryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for EXTERNAL services (e.g., AMS Disposal) to create Journal Entries.
    Protected by API Key.
    """
    queryset = JournalEntry.objects.all()
    # Use the new robust serializer
    serializer_class = ExternalJournalEntrySerializer 
    authentication_classes = [APIKeyAuthentication]
    permission_classes = [IsTrustedService]
    http_method_names = ['post']

    def perform_create(self, serializer):
        serializer.save()

# MODIFICATION START: Manage Budget Transfers (Approval Workflow)


@extend_schema(tags=['Supplemental Budget'])
class BudgetTransferViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet to List Pending Transfers and Approve/Reject them.
    Used for the 'Supplemental Budget Approval' tab in BudgetAllocation.jsx.
    """
    # MODIFIED: Allow all BMS users to view list, but restrict actions
    permission_classes = [IsBMSUser]
    serializer_class = BudgetTransferSerializer
    pagination_class = FiveResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = [
        'reason', 'destination_allocation__department__name', 'transferred_by_username']
    filterset_fields = ['status', 'transfer_type']

    def get_queryset(self):
        user = self.request.user
        bms_role = get_user_bms_role(user)

        # Base Query
        queryset = BudgetTransfer.objects.filter(
            transfer_type='SUPPLEMENTAL'
        ).select_related(
            'destination_allocation__department',
            'destination_allocation__category'
        ).order_by('-transferred_at')

        # DATA ISOLATION
        if bms_role in ['ADMIN', 'FINANCE_HEAD']:
            return queryset
        else:
            # General User (Operator) sees only their department's requests
            department_id = getattr(user, 'department_id', None)
            if department_id:
                return queryset.filter(destination_allocation__department_id=department_id)
            return queryset.none()

    @extend_schema(
        summary="Approve a Supplemental Budget Request",
        request=None,
        responses={200: OpenApiResponse(
            description="Request Approved and Budget Updated")}
    )
    # MODIFIED: Explicit permission check for action
    @extend_schema(
        summary="Approve a Supplemental Budget Request",
        request=None,
        responses={200: OpenApiResponse(
            description="Request Approved and Budget Updated")}
    )
    # MODIFIED: Explicit permission check for action
    @action(detail=True, methods=['post'], permission_classes=[IsBMSFinanceHead])
    def approve(self, request, pk=None):
        transfer = self.get_object()

        if transfer.status != 'PENDING':
            return Response(
                {"error": "Only pending requests can be approved."},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # 1. Get the allocation and verify department exists
            allocation = transfer.destination_allocation

            # --- FIX: Explicit Department Check ---
            if not allocation.department:
                return Response(
                    {"error": "Allocation is missing department. Cannot process supplemental budget."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            department = allocation.department

            # --- DEBUG LOGGING (Remove in production) ---
            print(f"🔍 Approving Transfer {transfer.id}")
            print(f"   Allocation ID: {allocation.id}")
            print(
                f"   Department ID: {department.id if department else 'NONE'}")
            print(
                f"   Department Name: {department.name if department else 'NONE'}")

            # 2. Find Equity/Treasury Account
            equity_account = Account.objects.filter(
                Q(account_type__name__iexact='Equity') |
                Q(name__icontains='Retained Earnings') |
                Q(name__icontains='Treasury')
            ).first()

            if not equity_account:
                return Response(
                    {"error": "Configuration Error: No 'Equity' or 'Treasury' account found."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # 3. Update Transfer Status
            transfer.status = 'APPROVED'
            transfer.approved_by_user_id = request.user.id
            transfer.approved_by_username = getattr(
                request.user, 'username', 'N/A')
            transfer.approval_date = timezone.now()
            transfer.save()

            # 4. Update Allocation (Add Money)
            allocation.amount += transfer.amount
            allocation.save()

            # 5. Create Journal Entry with EXPLICIT Department
            je = JournalEntry.objects.create(
                date=timezone.now().date(),
                category='PROJECTS',
                description=f"Approved Supplemental Budget: {transfer.reason}",
                total_amount=transfer.amount,
                status='POSTED',
                department=department,  # ✅ Now guaranteed to exist
                created_by_user_id=request.user.id,
                created_by_username=getattr(request.user, 'username', 'N/A')
            )

            # --- VERIFICATION LOG ---
            print(
                f"✅ Created JE {je.entry_id} with Department: {je.department.name if je.department else 'STILL NONE!'}")

            # 6. Create Journal Lines
            JournalEntryLine.objects.create(
                journal_entry=je,
                account=allocation.account,
                transaction_type='DEBIT',
                journal_transaction_type='TRANSFER',
                amount=transfer.amount,
                description=f"Supplemental increase for {allocation.category.name}",
                expense_category=allocation.category
            )

            JournalEntryLine.objects.create(
                journal_entry=je,
                account=equity_account,
                transaction_type='CREDIT',
                journal_transaction_type='TRANSFER',
                amount=transfer.amount,
                description=f"Funding source: {equity_account.name}",
                expense_category=None
            )

        return Response(
            {"message": "Supplemental budget approved and ledger updated."},
            status=status.HTTP_200_OK
        )


class FiscalYearViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for Fiscal Years.
    Used for the Fiscal Year Management dashboard.
    """
    queryset = FiscalYear.objects.all().order_by('-start_date')
    serializer_class = FiscalYearSerializer
    permission_classes = [IsBMSFinanceHead]  # Restricted to Finance Head/Admin

    @action(detail=True, methods=['post'])
    def set_status(self, request, pk=None):
        """
        Custom endpoint to change status (Open, Locked, Closed).
        Payload: { "status": "Locked" }
        """
        fiscal_year = self.get_object()
        status_action = request.data.get('status')  # Open, Locked, Closed

        if status_action == 'Open':
            # Ensure only one open year exists?
            # For now, just set flags. Business logic usually implies one active.
            fiscal_year.is_active = True
            fiscal_year.is_locked = False
        elif status_action == 'Locked':
            fiscal_year.is_locked = True
            # Keep active=True so we can still read data, just not edit
        elif status_action == 'Closed':
            fiscal_year.is_active = False
            fiscal_year.is_locked = True
        else:
            return Response({'error': 'Invalid status'}, status=400)

        fiscal_year.save()
        return Response(self.get_serializer(fiscal_year).data)


@extend_schema(
    tags=['Ledger View'],
    summary="Retrieve Journal Entry Details",
    description="Get full details of a journal entry including all debit/credit lines and audit info.",
    responses={200: JournalEntryDetailSerializer}
)
class JournalEntryDetailView(generics.RetrieveAPIView):
    queryset = JournalEntry.objects.all()
    serializer_class = JournalEntryDetailSerializer
    permission_classes = [IsBMSUser]
    lookup_field = 'entry_id'  # We will look up by "JE-2026-XXXX"

class ExternalReferenceViewSet(viewsets.ViewSet):
    """
    Read-only endpoints for External Systems to fetch Master Data.
    Protected by API Key.
    """
    authentication_classes = [APIKeyAuthentication]
    permission_classes = [IsTrustedService]

    @action(detail=False, methods=['get'])
    def departments(self, request):
        """Get valid Departments (ID, Name, Code)"""
        depts = Department.objects.filter(is_active=True).values('id', 'name', 'code')
        return Response(list(depts))

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """
        Get valid Expense Categories.
        Optional filter: ?department_code=IT
        """
        dept_code = request.query_params.get('department_code')
        qs = ExpenseCategory.objects.filter(is_active=True)

        # INTELLIGENT FILTERING
        if dept_code:
            # 1. Try to filter by SubCategoryBudgetCap )(Strongest Link)
            # This confirms the department actually has a policy/cap for this category
            
            # Find categories linked to this dept via Caps
            cap_linked_ids = SubCategoryBudgetCap.objects.filter(
                department__code__iexact=dept_code,
                is_active=True
            ).values_list('expense_category_id', flat=True)
            
            if cap_linked_ids.exists():
                qs = qs.filter(id__in=cap_linked_ids)
            else:
                # 2. Fallback: Prefix Matching (e.g. IT-HOST matches IT)
                # If no caps are seeded yet, fallback to naming convention
                qs = qs.filter(code__istartswith=dept_code)

        data = qs.values('id', 'name', 'code', 'classification')
        return Response(list(data))

    @action(detail=False, methods=['get'])
    def fiscal_years(self, request):
        """Get Active Fiscal Year"""

        fy = FiscalYear.objects.filter(is_active=True).values(
            'id', 'name', 'start_date', 'end_date'
        ).first()
        return Response(fy)


    @action(detail=False, methods=['get'])
    def accounts(self, request):
        """Get valid General Ledger Accounts for External Systems"""
        # Return ID, Name, Code, and Type so they can filter assets vs expenses
        accounts = Account.objects.filter(is_active=True).values(
            'id', 'name', 'code', 'account_type__name'
        )
        return Response(list(accounts))
    
    #URL: GET /api/external-references/budget_caps/?department_code=IT
    #Header: X-API-Key: <valid-key>
    
    @action(detail=False, methods=['get'])
    def budget_caps(self, request):
        """
        Get Budget Caps and Remaining Balances for a Department.
        Required param: ?department_code=IT
        
        Returns:
        {
            "fiscal_year": "FY 2026",
            "department": "IT",
            "department_cap": {
                "limit_amount": 3000000.00,
                "spent_amount": 1200000.00,
                "remaining_amount": 1800000.00,
                "cap_type": "SOFT",
                "cap_percentage": 15.0,
                "current_usage_percentage": 40.0
            },
            "category_caps": [
                {
                    "category_code": "IT-HOST",
                    "category_name": "Server Hosting",
                    "limit_amount": 600000.00,
                    "spent_amount": 550000.00,
                    "remaining_amount": 50000.00,
                    "cap_type": "HARD",
                    "cap_percentage": 20.0,
                    "current_usage_percentage": 91.7
                }
            ]
        }
        """
        dept_code = request.query_params.get('department_code')
        if not dept_code:
            return Response({"error": "department_code is required"}, status=400)

        # 1. Get Active Fiscal Year
        today = timezone.now().date()
        fiscal_year = FiscalYear.objects.filter(
            start_date__lte=today, end_date__gte=today, is_active=True
        ).first()
        
        if not fiscal_year:
            return Response({"error": "No active fiscal year"}, status=404)

        try:
            dept = Department.objects.get(code__iexact=dept_code)
        except Department.DoesNotExist:
            return Response({"error": "Invalid Department Code"}, status=404)

        # 2. Get Department Cap Info
        dept_cap_info = None
        try:
            d_cap = DepartmentBudgetCap.objects.get(
                department=dept, fiscal_year=fiscal_year, is_active=True
            )
            
            # Calculate total org budget for context
            total_org_allocations = BudgetAllocation.objects.filter(
                fiscal_year=fiscal_year, is_active=True
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Department's limit based on percentage of total
            dept_limit = total_org_allocations * (d_cap.percentage_of_total / 100)
            
            # Current spending for this department
            current_dept_spent = Expense.objects.filter(
                department=dept, 
                budget_allocation__fiscal_year=fiscal_year,
                status__in=['APPROVED', 'SUBMITTED']
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Calculate Usage Percentage
            dept_usage_pct = 0.0
            if dept_limit > 0:
                dept_usage_pct = (current_dept_spent / dept_limit) * 100
            
            dept_cap_info = {
                "limit_amount": round(dept_limit, 2),
                "spent_amount": round(current_dept_spent, 2),
                "remaining_amount": round(max(dept_limit - current_dept_spent, 0), 2),
                "cap_type": d_cap.cap_type,
                # NEW FIELDS
                "cap_percentage": float(d_cap.percentage_of_total),  # The policy limit (e.g., 15%)
                "current_usage_percentage": round(float(dept_usage_pct), 1)  # How much of that limit is used (e.g., 40%)
            }
        except DepartmentBudgetCap.DoesNotExist:
            pass

        # 3. Get Sub-Category Caps
        sub_caps = []
        cat_caps_qs = SubCategoryBudgetCap.objects.filter(
            department=dept, fiscal_year=fiscal_year, is_active=True
        ).select_related('expense_category')

        # We need total dept allocation for calculation
        dept_total_alloc = BudgetAllocation.objects.filter(
            department=dept, fiscal_year=fiscal_year, is_active=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        for c_cap in cat_caps_qs:
            # Category limit as percentage of department budget
            cat_limit = dept_total_alloc * (c_cap.percentage_of_department / 100)
            
            # Current spending for this category in this department
            current_cat_spent = Expense.objects.filter(
                department=dept,
                category=c_cap.expense_category,
                budget_allocation__fiscal_year=fiscal_year,
                status__in=['APPROVED', 'SUBMITTED']
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

            # Calculate Usage Percentage
            cat_usage_pct = 0.0
            if cat_limit > 0:
                cat_usage_pct = (current_cat_spent / cat_limit) * 100

            sub_caps.append({
                "category_code": c_cap.expense_category.code,
                "category_name": c_cap.expense_category.name,
                "limit_amount": round(cat_limit, 2),
                "spent_amount": round(current_cat_spent, 2),
                "remaining_amount": round(max(cat_limit - current_cat_spent, 0), 2),
                "cap_type": c_cap.cap_type,
                # NEW FIELDS
                "cap_percentage": float(c_cap.percentage_of_department),  # Policy percentage (e.g., 20%)
                "current_usage_percentage": round(float(cat_usage_pct), 1)  # Actual usage (e.g., 91.7%)
            })

        return Response({
            "fiscal_year": fiscal_year.name,
            "department": dept.code,
            "department_cap": dept_cap_info,
            "category_caps": sub_caps
        })
    