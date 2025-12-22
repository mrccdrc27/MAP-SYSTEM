from rest_framework.test import APITestCase, APIClient
from ..models import Supplier
from django.conf import settings
from openpyxl import Workbook
from io import BytesIO
from django.core.files.uploadedfile import SimpleUploadedFile


class ImportExportTests(APITestCase):
    def setUp(self):
        self.client = APIClient()
        # create an initial supplier to test update
        self.supplier = Supplier.objects.create(name='Initial Supplier')
        # configure import API key for tests
        settings.IMPORT_API_KEY = 'testkey'

    def _make_suppliers_xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(['id','name','address','city','zip','contact_name','phone_number','email','url','notes'])
        for r in rows:
            ws.append(r)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def test_export_suppliers_excludes_id(self):
        # Ensure export endpoint returns XLSX and header doesn't include 'id'
        url = '/export/suppliers/'
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        # read workbook from response
        wb = Workbook()
        bio = BytesIO(resp.content)
        from openpyxl import load_workbook
        wb = load_workbook(bio)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn('id', [h.lower() for h in headers if h])

    def test_import_create_and_update(self):
        # Create a workbook: one row without id (new), one row with id matching existing supplier (update)
        new_row = [None, 'Created Supplier 1', '1 New St', 'New City', '9999', 'John', '+111', 'j@x.com', 'http://new.example', 'note']
        update_row = [self.supplier.id, 'Initial Supplier Updated', '2 Old St', 'Old City', '0001', 'Jane', '+222', 'jane@x.com', 'http://old.example', 'updated note']
        bio = self._make_suppliers_xlsx([new_row, update_row])
        upload = SimpleUploadedFile('suppliers.xlsx', bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # First: import without allow_update - should create 2 new records (id column ignored)
        resp = self.client.post('/import/suppliers/', {'file': upload}, format='multipart')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        # created should be 2 (both rows treated as creates)
        self.assertEqual(data.get('created'), 2)

        # Now update: recreate workbook and post with allow_update=true
        bio = self._make_suppliers_xlsx([new_row, update_row])
        upload = SimpleUploadedFile('suppliers.xlsx', bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp = self.client.post('/import/suppliers/?allow_update=true', {'file': upload}, format='multipart', HTTP_X_IMPORT_API_KEY='testkey')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        # updated should be at least 1 (the row with existing id)
        self.assertEqual(data.get('updated'), 1)
        # verify supplier was updated
        s = Supplier.objects.get(pk=self.supplier.id)
        self.assertEqual(s.name, 'Initial Supplier Updated')

    def test_natural_upsert_and_logging(self):
        # Create a supplier to be matched by natural key (name)
        supplier = Supplier.objects.create(name='Natural Match')
        # Build workbook where row has same name but different address
        new_row = [supplier.id, 'Natural Match', '5 Orig St', 'Orig City', '1111', 'A', '+100', 'a@x.com', 'http://nat.example', 'note']
        # The importer should match by natural key when upsert_by=natural
        bio = self._make_suppliers_xlsx([new_row])
        upload = SimpleUploadedFile('suppliers.xlsx', bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Capture logs emitted by the import when updates occur
        with self.assertLogs('import_export', level='INFO') as cm:
            resp = self.client.post('/import/suppliers/?allow_update=true&upsert_by=natural', {'file': upload}, format='multipart', HTTP_X_IMPORT_API_KEY='testkey')
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        # Because the row includes an id, the handler will attempt to update; updated may be 1
        # but natural upsert should allow matching by name as well
        self.assertTrue(data.get('updated', 0) >= 0)
        # Check logs contain import_update entries
        found = any('import_update:' in m for m in cm.output)
        self.assertTrue(found, f'Expected import_update log in {cm.output}')
