from io import BytesIO
import datetime
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from ..services.depreciation_report import generate_depreciation_report
from ..services.asset_report import generate_asset_report
from ..services.activity_report import generate_activity_report, get_activity_summary

# No ticket resolution here — report will only include status fields per request


class DepreciationReportAPIView(APIView):
    """Return depreciation report as XLSX (default) or JSON.

    Query params:
      - depreciation_id: int to filter by a specific depreciation record
      - format=xlsx (default) to download an XLSX file
      - format=json to return JSON results
    """

    def get(self, request):
        dep = request.query_params.get('depreciation_id')
        fmt = request.query_params.get('format', '').lower()

        try:
            dep_id = int(dep) if dep is not None else None
        except ValueError:
            return Response({"detail": "Invalid depreciation_id"}, status=status.HTTP_400_BAD_REQUEST)

        rows = generate_depreciation_report(depreciation_id=dep_id)

        # Explicit JSON request returns JSON (backwards compatible)
        if fmt == 'json':
            return Response({'results': rows})

        # CSV not supported for this report; prefer XLSX
        if fmt == 'csv':
            return Response({'detail': 'CSV export is not supported. Use format=xlsx or omit format to download an XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        # Default: produce XLSX. Lazy import openpyxl and provide install guidance if missing.
        try:
            import openpyxl  # type: ignore
            from openpyxl.workbook import Workbook  # type: ignore
        except Exception:
            return Response({
                'detail': 'openpyxl is not installed in the running Python environment.',
                'install_instructions': (
                    'Add openpyxl to backend/assets/requirements.txt and rebuild the assets image,\n'
                    'for example: docker-compose -f docker-compose.dev.yml build --no-cache assets && '
                    'docker-compose -f docker-compose.dev.yml up -d assets'
                )
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        fieldnames = [
            'assetId', 'product', 'statusName', 'depreciationName', 'duration',
            'currency', 'minimumValue', 'purchaseCost', 'currentValue', 'depreciated',
            'monthlyDepreciation', 'monthsLeft'
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'DepreciationReport'

        # Header
        ws.append(fieldnames)

        for r in rows:
            row = []
            for f in fieldnames:
                v = r.get(f, '')
                if isinstance(v, float):
                    # monthlyDepreciation keep more precision
                    if f == 'monthlyDepreciation':
                        row.append(round(v, 6))
                    else:
                        row.append(round(v, 2))
                else:
                    row.append(v)
            ws.append(row)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"depreciation_report_{datetime.date.today().isoformat()}.xlsx"
        response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class AssetReportAPIView(APIView):
    """Return asset report as XLSX (default) or JSON.

    Query params:
      - status_id: int to filter by a specific status
      - category_id: int to filter by a specific category
      - supplier_id: int to filter by a specific supplier
      - location_id: int to filter by a specific location
      - product_id: int to filter by a specific product
      - manufacturer_id: int to filter by a specific manufacturer
      - columns: comma-separated list of column IDs to include
      - format=xlsx (default) to download an XLSX file
      - format=json to return JSON results
    """

    # Map frontend column IDs to backend field names
    COLUMN_MAPPING = {
        'asset_id': 'assetId',
        'asset_name': 'name',
        'purchase_date': 'purchaseDate',
        'purchase_cost': 'purchaseCost',
        'currency': 'currency',
        'order_number': 'orderNumber',
        'serial_number': 'serialNumber',
        'warranty_expiration': 'warrantyExpiration',
        'notes': 'notes',
        'product_data': 'product',
        'category_data': 'category',
        'manufacturer_data': 'manufacturer',
        'status_data': 'statusName',
        'supplier_data': 'supplier',
        'location_data': 'location',
        'depreciation_data': 'depreciation',
        'checked_out_to': 'checkedOutTo',
        'last_next_audit_date': 'auditDates',
        'picture_data': 'image',
        'created_at': 'createdAt',
        'updated_at': 'updatedAt',
    }

    # All available fieldnames for the report
    ALL_FIELDNAMES = [
        'assetId', 'name', 'product', 'category', 'statusName', 'supplier',
        'manufacturer', 'location', 'serialNumber', 'orderNumber',
        'purchaseDate', 'purchaseCost', 'warrantyExpiration', 'notes',
        'currency', 'depreciation', 'checkedOutTo', 'auditDates', 'image',
        'createdAt', 'updatedAt'
    ]

    def get(self, request):
        # Parse filter parameters
        status_id = request.query_params.get('status_id')
        category_id = request.query_params.get('category_id')
        supplier_id = request.query_params.get('supplier_id')
        location_id = request.query_params.get('location_id')
        product_id = request.query_params.get('product_id')
        manufacturer_id = request.query_params.get('manufacturer_id')
        columns_param = request.query_params.get('columns', '')
        # Use 'export_format' instead of 'format' to avoid conflict with DRF's format suffix
        fmt = request.query_params.get('export_format', '').lower()

        # Validate and convert filter IDs
        try:
            status_id = int(status_id) if status_id else None
            category_id = int(category_id) if category_id else None
            supplier_id = int(supplier_id) if supplier_id else None
            location_id = int(location_id) if location_id else None
            product_id = int(product_id) if product_id else None
            manufacturer_id = int(manufacturer_id) if manufacturer_id else None
        except ValueError:
            return Response({"detail": "Invalid filter parameter. IDs must be integers."}, status=status.HTTP_400_BAD_REQUEST)

        rows = generate_asset_report(
            status_id=status_id,
            category_id=category_id,
            supplier_id=supplier_id,
            location_id=location_id,
            product_id=product_id,
            manufacturer_id=manufacturer_id,
        )

        # Determine which columns to include
        if columns_param:
            column_ids = [c.strip() for c in columns_param.split(',') if c.strip()]
            fieldnames = []
            for col_id in column_ids:
                if col_id in self.COLUMN_MAPPING:
                    field = self.COLUMN_MAPPING[col_id]
                    if field not in fieldnames:
                        fieldnames.append(field)
            # If no valid columns found, use defaults
            if not fieldnames:
                fieldnames = self.ALL_FIELDNAMES[:14]  # Original default fields
        else:
            # Default fieldnames (original behavior)
            fieldnames = [
                'assetId', 'name', 'product', 'category', 'statusName', 'supplier',
                'manufacturer', 'location', 'serialNumber', 'orderNumber',
                'purchaseDate', 'purchaseCost', 'warrantyExpiration', 'notes'
            ]

        # JSON format
        if fmt == 'json':
            # Filter rows to only include selected columns
            filtered_rows = []
            for row in rows:
                filtered_row = {k: row.get(k, '') for k in fieldnames}
                filtered_rows.append(filtered_row)
            return Response({'results': filtered_rows, 'count': len(filtered_rows)})

        # CSV not supported
        if fmt == 'csv':
            return Response({'detail': 'CSV export is not supported. Use export_format=xlsx or omit export_format to download an XLSX file.'}, status=status.HTTP_400_BAD_REQUEST)

        # Default: XLSX export
        try:
            from openpyxl.workbook import Workbook
        except ImportError:
            return Response({
                'detail': 'openpyxl is not installed in the running Python environment.',
                'install_instructions': (
                    'Add openpyxl to backend/assets/requirements.txt and rebuild the assets image.'
                )
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        wb = Workbook()
        ws = wb.active
        ws.title = 'AssetReport'

        # Header row with human-readable names
        header_names = self._get_header_names(fieldnames)
        ws.append(header_names)

        for r in rows:
            row = []
            for f in fieldnames:
                v = r.get(f, '')
                if isinstance(v, float):
                    row.append(round(v, 2))
                else:
                    row.append(v)
            ws.append(row)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"asset_report_{datetime.date.today().isoformat()}.xlsx"
        response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _get_header_names(self, fieldnames):
        """Convert field names to human-readable header names."""
        header_map = {
            'assetId': 'Asset ID',
            'name': 'Asset Name',
            'product': 'Product',
            'category': 'Category',
            'statusName': 'Status',
            'supplier': 'Supplier',
            'manufacturer': 'Manufacturer',
            'location': 'Location',
            'serialNumber': 'Serial Number',
            'orderNumber': 'Order Number',
            'purchaseDate': 'Purchase Date',
            'purchaseCost': 'Purchase Cost',
            'warrantyExpiration': 'Warranty Expiration',
            'notes': 'Notes',
            'currency': 'Currency',
            'depreciation': 'Depreciation',
            'checkedOutTo': 'Checked Out To',
            'auditDates': 'Audit Dates',
            'image': 'Image',
            'createdAt': 'Created At',
            'updatedAt': 'Updated At',
        }
        return [header_map.get(f, f) for f in fieldnames]


class ActivityReportAPIView(APIView):
    """Return activity report as XLSX (default) or JSON.

    Query params:
      - start_date: Filter activities from this date (YYYY-MM-DD format)
      - end_date: Filter activities up to this date (YYYY-MM-DD format)
      - activity_type: Filter by activity type (Asset, Component, Audit, Repair)
      - action: Filter by action (Create, Update, Delete, Checkout, Checkin, etc.)
      - user_id: Filter by the user who performed the action
      - item_id: Filter by the item ID affected
      - search: Search term to filter by item name/identifier, user name, or notes
      - limit: Maximum number of records to return
      - export_format: 'xlsx' (default) or 'json'
    """

    # Fieldnames for the report
    FIELDNAMES = ['date', 'user', 'type', 'action', 'item', 'to_from', 'notes']

    def get(self, request):
        # Parse filter parameters
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        activity_type = request.query_params.get('activity_type')
        action = request.query_params.get('action')
        user_id_param = request.query_params.get('user_id')
        item_id_param = request.query_params.get('item_id')
        search = request.query_params.get('search')
        limit_param = request.query_params.get('limit')
        fmt = request.query_params.get('export_format', '').lower()

        # Parse and validate date parameters
        start_date = None
        end_date = None
        try:
            if start_date_str:
                start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"detail": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate and convert integer parameters
        try:
            user_id = int(user_id_param) if user_id_param else None
            item_id = int(item_id_param) if item_id_param else None
            limit = int(limit_param) if limit_param else None
        except ValueError:
            return Response(
                {"detail": "Invalid parameter. user_id, item_id, and limit must be integers."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Generate the report data
        rows = generate_activity_report(
            start_date=start_date,
            end_date=end_date,
            activity_type=activity_type,
            action=action,
            user_id=user_id,
            item_id=item_id,
            search=search,
            limit=limit,
        )

        # JSON format
        if fmt == 'json':
            return Response({
                'results': rows,
                'count': len(rows),
            })

        # CSV not supported
        if fmt == 'csv':
            return Response(
                {'detail': 'CSV export is not supported. Use export_format=xlsx or omit export_format to download an XLSX file.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Default: XLSX export
        try:
            from openpyxl.workbook import Workbook
        except ImportError:
            return Response({
                'detail': 'openpyxl is not installed in the running Python environment.',
                'install_instructions': (
                    'Add openpyxl to backend/assets/requirements.txt and rebuild the assets image.'
                )
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        wb = Workbook()
        ws = wb.active
        ws.title = 'ActivityReport'

        # Header row
        header_names = ['Date', 'User', 'Type', 'Event', 'Item', 'To/From', 'Notes']
        ws.append(header_names)

        for r in rows:
            row = [
                r.get('date', ''),
                r.get('user', ''),
                r.get('type', ''),
                r.get('action', ''),
                r.get('item', ''),
                r.get('to_from', ''),
                r.get('notes', ''),
            ]
            ws.append(row)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"activity_report_{datetime.date.today().isoformat()}.xlsx"
        response = HttpResponse(
            bio.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ActivityReportSummaryAPIView(APIView):
    """Return activity report summary statistics.

    Query params:
      - start_date: Filter activities from this date (YYYY-MM-DD format)
      - end_date: Filter activities up to this date (YYYY-MM-DD format)
    """

    def get(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        # Parse and validate date parameters
        start_date = None
        end_date = None
        try:
            if start_date_str:
                start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"detail": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )

        summary = get_activity_summary(start_date=start_date, end_date=end_date)
        return Response(summary)
